#!/usr/bin/env python3
"""Low-frequency collectors for the configured recruitment sources.

Every source is requested at most a few times per run.  The dedicated adapters
below cover static government columns, seasonal civil-service pages and the
public data embedded by large-company recruitment sites.
"""

from __future__ import annotations

import hashlib
import io
import json
import re
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Callable
from urllib.parse import quote_plus, urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCES_PATH = ROOT / "data" / "sources.json"
OUTPUT_PATH = ROOT / "data" / "collected" / "other-sources.json"
BEIJING_TZ = timezone(timedelta(hours=8))
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 BJJob/0.2"
INCLUDE_RE = re.compile(r"招聘|招录|考录|公务员|事业单位|校园招聘|社会招聘|实习|职位|岗位|选调|优培|补录|调剂")
EXCLUDE_RE = re.compile(r"登录|注册|隐私|关于我们|网站地图|联系我们|帮助|政策法规|成绩查询|报名入口")
DATE_RE = re.compile(r"(20\d{2})[-年./](\d{1,2})[-月./](\d{1,2})日?")
ATTACHMENT_RE = re.compile(r"\.(xlsx?|docx?|pdf|zip|rar)(?:\?.*)?$", re.I)
WORKBOOK_FIELDS = {
    "organization": ("招聘单位", "单位名称", "用人单位", "所属单位"),
    "title": ("岗位名称", "招聘岗位", "职位名称", "岗位"),
    "headcount": ("招聘人数", "拟招聘人数", "计划人数"),
    "education": ("学历要求", "学历", "最低学历"),
    "degree": ("学位要求", "学位"),
    "major": ("专业要求", "所学专业", "专业"),
    "applicant_type": ("招聘对象", "人员类别", "应聘人员类别"),
    "requirements": ("其他条件", "其它条件", "资格条件", "岗位要求", "任职要求"),
    "responsibilities": ("岗位职责", "工作职责", "主要职责", "职责描述"),
}


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def now_iso() -> str:
    return datetime.now(BEIJING_TZ).replace(microsecond=0).isoformat()


def make_item(source: dict[str, Any], title: str, url: str, published: str = "", organization: str = "",
              **details: Any) -> dict[str, Any]:
    item = {
        "id": hashlib.sha1(url.encode("utf-8")).hexdigest()[:16],
        "title": clean(title)[:240],
        "organization": clean(organization) or source["name"],
        "category": source["group"],
        "published_at": published,
        "source_name": source["name"],
        "source_url": url,
        "source_home": source["url"],
    }
    item.update({key: clean(value) for key, value in details.items() if value not in (None, "", [], {})})
    return item


def published_from(text: str) -> str:
    match = DATE_RE.search(text)
    if not match:
        return ""
    year, month, day = match.groups()
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def same_site(base: str, candidate: str) -> bool:
    a = urlparse(base).netloc.removeprefix("www.")
    b = urlparse(candidate).netloc.removeprefix("www.")
    return b == a or b.endswith("." + a)


def canonical_workbook_field(value: Any) -> str | None:
    header = re.sub(r"\s+", "", clean(value))
    exact_only = {"岗位", "专业", "学历", "学位"}
    for field, aliases in WORKBOOK_FIELDS.items():
        if any(
            header == alias or (alias not in exact_only and alias in header and len(header) <= len(alias) + 5)
            for alias in aliases
        ):
            return field
    return None


def parse_links(source: dict[str, Any], response: requests.Response, allow_external: bool = False,
                href_pattern: str | None = None) -> list[dict[str, str]]:
    soup = BeautifulSoup(response.text, "html.parser")
    found: dict[str, dict[str, str]] = {}
    for link in soup.find_all("a", href=True):
        title = clean(link.get("title") or link.get_text(" ", strip=True))
        href = clean(link.get("href"))
        wrapped = re.match(r"javascript:checkUrl\(['\"](.+?)['\"]\)", href)
        if wrapped:
            href = wrapped.group(1)
        if len(title) < 4 or EXCLUDE_RE.search(title):
            continue
        if not INCLUDE_RE.search(title) and not (href_pattern and re.search(href_pattern, href)):
            continue
        url = urljoin(response.url, href)
        if not url.startswith(("http://", "https://")):
            continue
        if not allow_external and not same_site(response.url, url):
            continue
        context = clean(link.parent.get_text(" ", strip=True)) if link.parent else title
        found[url] = make_item(source, title, url, published_from(context))
    return list(found.values())[:100]


def extract_detail(session: requests.Session, item: dict[str, Any]) -> dict[str, Any]:
    """Read one announcement body and extract conditions, dates and attachments."""
    response = session.get(item["source_url"], timeout=30, headers={"Referer": item["source_home"]})
    response.raise_for_status()
    response.encoding = response.apparent_encoding or "utf-8"
    soup = BeautifulSoup(response.text, "html.parser")
    for node in soup.select("script,style,nav,header,footer"):
        node.decompose()
    container = soup.select_one("article,.TRS_Editor,.article-content,.content,#zoom,#content") or soup.body or soup
    text = clean(container.get_text(" ", strip=True))
    if len(text) < 30:
        return item

    # Capture the most useful section without storing an entire mirrored article.
    condition = ""
    match = re.search(r"(?:招聘|报考|应聘|资格)(?:对象|范围|条件|要求)[：:]?(.{20,1600}?)(?=报名|招聘程序|考试|考核|薪酬|联系方式|附件|$)", text)
    if match:
        condition = clean(match.group(1))
    elif len(text) <= 1800:
        condition = text

    deadline = ""
    date_candidates = DATE_RE.findall(" ".join(re.findall(r"[^。；]{0,20}(?:报名|申请|投递|截止)[^。；]{0,80}", text)))
    if date_candidates:
        year, month, day = date_candidates[-1]
        deadline = f"{int(year):04d}-{int(month):02d}-{int(day):02d}T23:59:00+08:00"

    attachments = []
    for link in soup.find_all("a", href=True):
        url = urljoin(response.url, link["href"])
        if ATTACHMENT_RE.search(url):
            attachments.append({"name": clean(link.get_text(" ", strip=True)) or Path(urlparse(url).path).name, "url": url})
    item.update({"requirements": condition, "deadline": deadline, "detail_parsed": True,
                 "attachments": attachments[:20], "data_quality": "已读取公告正文"})
    return item


def workbook_positions(content: bytes, attachment_url: str, notice: dict[str, Any]) -> list[dict[str, Any]]:
    """Convert a conventional government position workbook into individual jobs."""
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    result = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        header_index, header_map = -1, {}
        for index, row in enumerate(rows[:20]):
            mapped = {}
            for column, value in enumerate(row):
                field = canonical_workbook_field(value)
                if field:
                    mapped[column] = field
            if len(set(mapped.values())) >= 3 and ("title" in mapped.values() or "organization" in mapped.values()):
                header_index, header_map = index, mapped
                break
        if header_index < 0:
            continue
        for row_number, row in enumerate(rows[header_index + 1:], start=header_index + 2):
            values = {header_map[column]: clean(value) for column, value in enumerate(row) if column in header_map and clean(value)}
            if not values.get("title") and not values.get("organization"):
                continue
            url = f"{notice['source_url']}#position-{sheet.title}-{row_number}"
            result.append(make_item(
                {"name": notice["source_name"], "group": notice["category"], "url": notice["source_home"]},
                values.get("title") or notice["title"], url, notice.get("published_at", ""),
                values.get("organization") or notice.get("organization", ""),
                **{key: value for key, value in values.items() if key not in ("title", "organization")},
                deadline=notice.get("deadline"), source_attachment_url=attachment_url,
                recruitment_type="公告附件岗位", data_quality="附件岗位表已解析",
            ))
    return result[:500]


def static_adapter(session: requests.Session, source: dict[str, Any], **kwargs: Any) -> tuple[list[dict[str, str]], str, str]:
    response = session.get(source["url"], timeout=30, allow_redirects=True)
    response.raise_for_status()
    response.encoding = response.apparent_encoding or "utf-8"
    items = parse_links(source, response, **kwargs)
    if source["group"] != "互联网大厂":
        for index, item in enumerate(items[:10]):
            if not same_site(response.url, item["source_url"]):
                continue
            try:
                if index: time.sleep(0.5)
                extract_detail(session, item)
            except Exception as exc:
                item["detail_error"] = clean(exc)[:160]
        attachment_budget = 3
        positions = []
        for item in items[:10]:
            for attachment in item.get("attachments", []):
                if attachment_budget <= 0 or not re.search(r"\.xlsx?(?:\?.*)?$", attachment["url"], re.I):
                    continue
                try:
                    attachment_budget -= 1
                    content = session.get(attachment["url"], timeout=40, headers={"Referer": item["source_url"]}).content
                    parsed = workbook_positions(content, attachment["url"], item)
                    attachment["position_count"] = len(parsed)
                    positions.extend(parsed)
                except Exception as exc:
                    attachment["parse_error"] = clean(exc)[:160]
        if positions:
            items = positions + items
    return items, "collected" if items else "collected-empty", response.url


def yearly_civil_service(session: requests.Session, source: dict[str, Any]) -> tuple[list[dict[str, str]], str, str]:
    # The annual host is only populated during recruitment. Probe current and next
    # annual专题; an inactive redirect is a valid seasonal state, not an error.
    years = sorted({datetime.now(BEIJING_TZ).year, datetime.now(BEIJING_TZ).year + 1}, reverse=True)
    for year in years:
        url = f"http://bm.scs.gov.cn/kl{year}"
        response = session.get(url, timeout=30, allow_redirects=True)
        response.encoding = response.apparent_encoding or "utf-8"
        if response.ok and len(response.text) > 1000:
            current = {**source, "url": url, "name": f"{year}年度国考专题"}
            items = parse_links(current, response, allow_external=True)
            if items:
                return items, "collected", response.url
    return [], "seasonal-inactive", response.url


def mohrss_challenge_cookies(html: str) -> dict[str, str]:
    """Decode the small arithmetic cookie challenge used by www.mohrss.gov.cn."""
    if "EO_Bot_Ssid" not in html:
        return {}
    status = re.search(r"WTKkN:(\d+),bOYDu:(\d+).*?wyeCN:(\d+)", html)
    session_id = re.search(r"\(t,(\d{8,13})\)", html)
    if not status or not session_id:
        raise RuntimeError("Unsupported mohrss.gov.cn cookie challenge")
    return {
        "__tst_status": f"{sum(int(value) for value in status.groups())}#",
        "EO_Bot_Ssid": session_id.group(1),
    }


def mohrss_adapter(session: requests.Session, source: dict[str, Any]) -> tuple[list[dict[str, str]], str, str]:
    """Collect the two recruitment columns after satisfying the site's cookie challenge."""
    response = session.get(source["url"], timeout=30, allow_redirects=True)
    response.raise_for_status()
    cookies = mohrss_challenge_cookies(response.text)
    if cookies:
        domain = urlparse(response.url).hostname
        for name, value in cookies.items():
            session.cookies.set(name, value, domain=domain)

    found: dict[str, dict[str, Any]] = {}
    final_url = response.url
    for section in ("zpgg/", "gxbyszpzl/"):
        section_source = {**source, "url": urljoin(source["url"], section)}
        items, _, final_url = static_adapter(session, section_source)
        for item in items:
            article = re.search(r"/t\d+_(\d+)\.html(?:#(.+))?$", item["source_url"])
            if not article:
                continue
            key = f"{article.group(1)}#{article.group(2) or ''}"
            found.setdefault(key, item)

    values = list(found.values())
    return values, "collected" if values else "collected-empty", final_url


def baidu_adapter(session: requests.Session, source: dict[str, Any]) -> tuple[list[dict[str, str]], str, str]:
    found: dict[str, dict[str, str]] = {}
    final_url = source["url"]
    for recruit_type in ("GRADUATE", "INTERN", "SOCIAL"):
        url = f"https://talent.baidu.com/jobs/list?recruitType={recruit_type}"
        response = session.get(url, timeout=40)
        response.raise_for_status()
        final_url = response.url
        marker = re.search(r"window\.__INITIAL_DATA__\s*=\s*", response.text)
        if not marker:
            continue
        try:
            # Baidu serializes JavaScript `undefined` in this otherwise-JSON object.
            serialised = response.text[marker.end():].replace("undefined", "null")
            payload, _ = json.JSONDecoder().raw_decode(serialised)
        except json.JSONDecodeError:
            continue

        def walk(value: Any) -> None:
            if isinstance(value, dict):
                title = clean(value.get("name") or value.get("jobName") or value.get("title"))
                place = clean(value.get("workPlace") or value.get("workplace") or value.get("city"))
                code = clean(value.get("jobId") or value.get("postId") or value.get("jobCode") or value.get("code") or value.get("id"))
                is_job = bool(value.get("serviceCondition") or value.get("workContent"))
                if title and code and is_job and ("北京" in place or not place):
                    detail = f"https://talent.baidu.com/jobs/detail/{recruit_type}/{code}"
                    requirements = clean(value.get("serviceCondition"))
                    education_match = re.search(r"(博士|硕士|本科|大专)(?:研究生)?(?:及以上)?", requirements)
                    found[detail] = make_item(
                        source, title, detail, clean(value.get("publishDate")), "百度",
                        requirements=requirements,
                        responsibilities=value.get("workContent"),
                        location=place,
                        education=education_match.group(0) if education_match else value.get("education"),
                        category_detail=value.get("postType"),
                        recruitment_type={"GRADUATE": "校园招聘", "INTERN": "实习", "SOCIAL": "社会招聘"}.get(recruit_type, recruit_type),
                        headcount=value.get("recruitNum"),
                        updated_at=value.get("updateDate"),
                        data_quality="岗位详情完整",
                    )
                for child in value.values():
                    walk(child)
            elif isinstance(value, list):
                for child in value:
                    walk(child)
        walk(payload)
    values = list(found.values())[:200]
    return values, "collected" if values else "collected-empty", final_url


def jd_adapter(session: requests.Session, source: dict[str, Any]) -> tuple[list[dict[str, str]], str, str]:
    endpoint = "https://zhaopin.jd.com/web/job/job_list"
    response = session.post(endpoint, timeout=40, headers={"Referer": source["url"]}, data={
        "pageIndex": 1, "pageSize": 100, "workCityJson": '["11"]',
        "jobTypeJson": "[]", "jobSearch": "", "depTypeJson": "[]",
    })
    response.raise_for_status()
    rows = response.json()
    found = []
    for row in rows:
        if "北京" not in clean(row.get("workCity")):
            continue
        title = clean(row.get("positionNameOpen") or row.get("positionName"))
        ident = clean(row.get("requirementId") or row.get("positionId"))
        if not title or not ident:
            continue
        requirements = clean(row.get("qualification"))
        education_match = re.search(r"(博士|硕士|本科|大专)(?:研究生)?(?:及以上)?", requirements)
        detail = f"https://zhaopin.jd.com/web/job/job_info_list/3?jobSearch={quote_plus(title)}#job-{ident}"
        found.append(make_item(
            source, title, detail, clean(row.get("formatPublishTime")),
            clean(row.get("positionDeptName")) or "京东",
            requirements=requirements, responsibilities=row.get("workContent"),
            location=row.get("workCity"), education=education_match.group(0) if education_match else "",
            category_detail=row.get("jobType"), recruitment_type="社会招聘",
            position_code=row.get("positionCode"), data_quality="岗位接口详情完整",
        ))
    return found, "collected" if found else "collected-empty", endpoint


def bytedance_adapter(session: requests.Session, source: dict[str, Any]) -> tuple[list[dict[str, str]], str, str]:
    """Do not expose search-indexed detail URLs without a live-job signal."""
    response = session.get(source["url"], timeout=30)
    response.raise_for_status()
    return [], "adapter-blocked", response.url


def tencent_adapter(session: requests.Session, source: dict[str, Any]) -> tuple[list[dict[str, str]], str, str]:
    endpoint = "https://careers.tencent.com/tencentcareer/api/post/Query"
    found = []
    verified_at = now_iso()
    for page_index in range(1, 4):
        response = session.get(endpoint, timeout=40, params={
            "timestamp": int(time.time() * 1000), "countryId": "", "cityId": "2",
            "bgIds": "", "productId": "", "categoryId": "", "parentCategoryId": "",
            "attrId": "", "keyword": "", "pageIndex": page_index, "pageSize": 100,
            "language": "zh-cn", "area": "cn",
        }, headers={"Referer": source["url"]})
        response.raise_for_status()
        rows = ((response.json().get("Data") or {}).get("Posts") or [])
        for row in rows:
            if clean(row.get("LocationName")) != "北京" or not row.get("IsValid", True):
                continue
            position_id = clean(row.get("PostId"))
            title = clean(row.get("RecruitPostName"))
            if not position_id or not title:
                continue
            requirements = clean(row.get("Responsibility"))
            experience = clean(row.get("RequireWorkYearsName"))
            found.append(make_item(
                source, title, f"https://careers.tencent.com/jobdesc.html?postId={position_id}",
                published_from(clean(row.get("LastUpdateTime"))), "腾讯",
                location="北京", requirements=" ".join(filter(None, [experience, requirements])),
                category_detail=row.get("CategoryName"), department=row.get("BGName"),
                product=row.get("ProductName"), recruitment_type="社会招聘",
                last_verified_at=verified_at, data_quality="腾讯官方职位接口",
            ))
        if len(rows) < 100:
            break
    return found, "collected" if found else "collected-empty", endpoint


XIAOMI_DESIGN_POSITIONS = ["886", "887", "884-1254", "905", "951"]


def xiaomi_adapter(session: requests.Session, source: dict[str, Any]) -> tuple[list[dict[str, str]], str, str]:
    found = []
    verified_at = now_iso()
    for position_id in XIAOMI_DESIGN_POSITIONS:
        url = f"https://hr.xiaomi.com/campus/view/{position_id}"
        response = session.get(url, timeout=30, headers={"Referer": source["url"]})
        if not response.ok:
            continue
        response.encoding = response.apparent_encoding or "utf-8"
        soup = BeautifulSoup(response.text, "html.parser")
        page_title = clean(soup.title.get_text(" ", strip=True) if soup.title else "")
        match = re.match(r"小米-北京-(.+?)-(.+?)-职位详情", page_title)
        if not match:
            continue
        category, title = match.groups()
        body = clean(soup.get_text(" ", strip=True))
        duty = re.search(r"工作职责[：:]?(.{10,1600}?)(?=工作要求|申请职位|$)", body)
        requirement = re.search(r"工作要求[：:]?(.{10,1600}?)(?=申请职位|热门|$)", body)
        found.append(make_item(
            source, title, url, organization="小米", location="北京",
            requirements=duty.group(1) if duty else "",
            responsibilities=requirement.group(1) if requirement else "",
            category_detail=category, recruitment_type="校园招聘/实习",
            last_verified_at=verified_at, data_quality="小米官方职位详情页",
        ))
        time.sleep(0.15)
    return found, "collected" if found else "collected-empty", source["url"]


def api_spa_adapter(session: requests.Session, source: dict[str, Any], endpoint: str) -> tuple[list[dict[str, str]], str, str]:
    """Probe a documented public SPA endpoint; keep failures distinct from empty data."""
    response = session.get(endpoint, timeout=30, headers={"Referer": source["url"]})
    if not response.ok:
        response.raise_for_status()
    content_type = response.headers.get("Content-Type", "")
    if "json" not in content_type:
        return [], "adapter-blocked", response.url
    data = response.json()
    found: dict[str, dict[str, str]] = {}
    def walk(value: Any) -> None:
        if isinstance(value, dict):
            title = clean(value.get("jobName") or value.get("name") or value.get("title"))
            city = clean(value.get("workCity") or value.get("cityName") or value.get("workPlace"))
            ident = clean(value.get("jobId") or value.get("id") or value.get("code"))
            if title and ident and (not city or "北京" in city):
                url = source["url"].split("?")[0] + ("&" if "?" in source["url"] else "?") + "jobId=" + ident
                found[url] = make_item(source, title, url)
            for child in value.values(): walk(child)
        elif isinstance(value, list):
            for child in value: walk(child)
    walk(data)
    values = list(found.values())[:200]
    return values, "collected" if values else "collected-empty", response.url


SPECIAL: dict[str, Callable[..., tuple[list[dict[str, str]], str, str]]] = {
    "bj-civil-service": lambda s, x: static_adapter(s, x, allow_external=True),
    "bj-exam-notices": lambda s, x: static_adapter(s, x),
    "national-civil-service-yearly": yearly_civil_service,
    "mohrss-central-institutions": mohrss_adapter,
    "bj-sasac-jobs": lambda s, x: static_adapter(s, x, allow_external=True),
    "iguopin": lambda s, x: api_spa_adapter(s, x, "https://www.iguopin.com/api/jobs/v3/list"),
    "ggj-notices": lambda s, x: static_adapter(s, x),
    "cnipa-personnel": lambda s, x: static_adapter(s, x, href_pattern=r"/art/\d{4}/.*art_74_"),
    "bytedance-jobs": bytedance_adapter,
    "baidu-jobs": baidu_adapter,
    "jd-jobs": jd_adapter,
    "meituan-jobs": lambda s, x: api_spa_adapter(s, x, "https://zhaopin.meituan.com/api/official/job/getJobList"),
    "tencent-jobs": tencent_adapter,
    "xiaomi-jobs": xiaomi_adapter,
}


def collect_source(session: requests.Session, source: dict[str, Any]) -> dict[str, Any]:
    started = time.monotonic()
    try:
        adapter = SPECIAL.get(source["id"], static_adapter)
        items, status, final_url = adapter(session, source)
        return {"source_id": source["id"], "source_name": source["name"], "group": source["group"],
                "home": source["url"], "adapter": "dedicated" if source["id"] in SPECIAL else "static",
                "status": status, "final_url": final_url, "item_count": len(items),
                "duration_ms": round((time.monotonic() - started) * 1000), "items": items}
    except Exception as exc:
        status_code = getattr(getattr(exc, "response", None), "status_code", None)
        return {"source_id": source["id"], "source_name": source["name"], "group": source["group"],
                "home": source["url"], "adapter": "dedicated" if source["id"] in SPECIAL else "static",
                "status": "adapter-blocked" if source["id"] in SPECIAL else "unavailable",
                "http_status": status_code, "item_count": 0, "error": clean(exc)[:300], "items": []}


def main() -> int:
    sources = [s for s in json.loads(SOURCES_PATH.read_text(encoding="utf-8")) if s["id"] != "bj-rsj-institutions"]
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT, "Accept-Language": "zh-CN,zh;q=0.9", "Accept": "text/html,application/json"})
    reports = []
    for index, source in enumerate(sources):
        if index: time.sleep(2)
        report = collect_source(session, source)
        reports.append(report)
        print(f"{source['id']}: {report['status']} ({report['item_count']})")
    all_items = {item["source_url"]: item for report in reports for item in report["items"]}
    output = {"generated_at": now_iso(), "source_count": len(reports),
              "collected_source_count": sum(r["status"] in ("collected", "collected-empty", "seasonal-inactive") for r in reports),
              "needs_adapter_count": 0,
              "unavailable_count": sum(r["status"] in ("unavailable", "adapter-blocked") for r in reports),
              "item_count": len(all_items), "sources": reports,
              "items": sorted(all_items.values(), key=lambda i: (i["published_at"], i["title"]), reverse=True)}
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(output, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
