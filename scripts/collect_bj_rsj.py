#!/usr/bin/env python3
"""Incrementally collect Beijing public-institution recruitment notices.

The collector is deliberately conservative: it reads the newest listing page,
fetches only new or changed notices, sleeps between requests, and keeps source
URLs for every notice, attachment, and extracted position.
"""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import random
import re
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


BASE_URL = "https://rsj.beijing.gov.cn"
LIST_URL = f"{BASE_URL}/xxgk/gkzp/"
ROOT = Path(__file__).resolve().parents[1]
STATE_PATH = ROOT / "data" / "collector-state" / "bj-rsj.json"
OUTPUT_PATH = ROOT / "data" / "collected" / "bj-rsj.json"
BEIJING_TZ = timezone(timedelta(hours=8))
USER_AGENT = "BJJob/0.1 (personal low-frequency recruitment index; source links preserved)"

NOTICE_LINK_RE = re.compile(r"/xxgk/gkzp/\d{6}/t\d+_\d+\.html$")
ATTACHMENT_RE = re.compile(r"\.(xlsx?|docx?|pdf|zip|rar)(?:\?.*)?$", re.I)
DATE_RE = re.compile(r"(20\d{2})[-年/.](\d{1,2})[-月/.](\d{1,2})日?")

FIELD_ALIASES = {
    "organization": ("招聘单位", "单位名称", "用人单位", "所属单位"),
    "title": ("岗位名称", "招聘岗位", "职位名称"),
    "category": ("岗位类别", "职位类别", "岗位类型"),
    "headcount": ("招聘人数", "拟招聘人数", "计划人数"),
    "education": ("学历要求", "学历", "最低学历"),
    "degree": ("学位要求", "学位"),
    "major": ("专业要求", "所学专业", "专业"),
    "age": ("年龄要求", "年龄"),
    "household": ("户籍要求", "生源要求", "户口要求"),
    "applicant_type": ("招聘对象", "人员类别", "应聘人员类别"),
    "requirements": ("其他条件", "其它条件", "资格条件", "岗位要求"),
    "contact": ("联系电话", "咨询电话", "联系方式"),
}


def now_iso() -> str:
    return datetime.now(BEIJING_TZ).replace(microsecond=0).isoformat()


def clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


class Collector:
    def __init__(self, delay_min: float, delay_max: float, timeout: int = 30):
        self.delay_min = delay_min
        self.delay_max = delay_max
        self.timeout = timeout
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": USER_AGENT, "Accept-Language": "zh-CN,zh;q=0.9"})
        self.last_request_at = 0.0

    def _wait(self) -> None:
        if not self.last_request_at:
            return
        target = random.uniform(self.delay_min, self.delay_max)
        elapsed = time.monotonic() - self.last_request_at
        if elapsed < target:
            time.sleep(target - elapsed)

    def get(self, url: str) -> requests.Response:
        self._wait()
        response = self.session.get(url, timeout=self.timeout)
        self.last_request_at = time.monotonic()
        response.raise_for_status()
        return response


def parse_listing(html: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    found: dict[str, dict[str, str]] = {}
    for link in soup.find_all("a", href=True):
        url = urljoin(LIST_URL, link["href"])
        if not NOTICE_LINK_RE.search(url):
            continue
        title = clean(link.get("title") or link.get_text(" ", strip=True))
        if not title:
            continue
        container_text = clean(link.parent.get_text(" ", strip=True)) if link.parent else title
        match = DATE_RE.search(container_text)
        published = "-".join(f"{int(part):02d}" for part in match.groups()) if match else ""
        found[url] = {"title": title, "url": url, "published_at": published}
    return list(found.values())


def content_hash(text: str) -> str:
    normalized = re.sub(r"\s+", "", text)
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def extract_period(text: str) -> tuple[str, str]:
    pattern = re.compile(
        r"(20\d{2})年(\d{1,2})月(\d{1,2})日"
        r"(?:\s*(\d{1,2})[:：时](\d{1,2})?分?)?[^。；]{0,20}?(?:至|—|-|到)"
        r"(?:(20\d{2})年)?(\d{1,2})月(\d{1,2})日"
        r"(?:\s*(\d{1,2})[:：时](\d{1,2})?分?)?"
    )
    candidate_sentences = [
        sentence for sentence in re.split(r"[。；]", text)
        if re.search(r"报名|报考|申请|提交应聘", sentence) and re.search(r"至|—|到", sentence)
    ]
    candidate_sentences.sort(
        key=lambda sentence: (
            0 if re.search(r"报名|提交应聘|申请时间|报名时间", sentence) else 1,
            len(sentence),
        )
    )
    match = next((pattern.search(sentence) for sentence in candidate_sentences if pattern.search(sentence)), None)
    if not match:
        return "", ""
    y1, m1, d1, h1, min1, y2, m2, d2, h2, min2 = match.groups()
    y2 = y2 or y1
    start = f"{int(y1):04d}-{int(m1):02d}-{int(d1):02d}T{int(h1 or 0):02d}:{int(min1 or 0):02d}:00+08:00"
    end = f"{int(y2):04d}-{int(m2):02d}-{int(d2):02d}T{int(h2 or 23):02d}:{int(min2 or 59):02d}:00+08:00"
    return start, end


def canonical_key(header: str) -> str | None:
    compact = re.sub(r"\s+", "", header)
    for key, aliases in FIELD_ALIASES.items():
        if any(compact == alias or (compact.endswith(alias) and len(compact) <= len(alias) + 4) for alias in aliases):
            return key
    return None


def parse_workbook(content: bytes, source_url: str) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    positions: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        header_index = -1
        header_map: dict[int, str] = {}
        for index, row in enumerate(rows[:20]):
            candidate = {column: canonical_key(clean(value)) for column, value in enumerate(row)}
            recognized = {column: key for column, key in candidate.items() if key}
            if len(recognized) >= 3 and ("title" in recognized.values() or "organization" in recognized.values()):
                header_index, header_map = index, recognized
                break
        if header_index < 0:
            continue
        raw_headers = {column: clean(value) for column, value in enumerate(rows[header_index]) if clean(value)}
        for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            values = {column: clean(value) for column, value in enumerate(row)}
            if not any(values.values()):
                continue
            item = {key: values.get(column, "") for column, key in header_map.items()}
            if not item.get("title") and not item.get("organization"):
                continue
            item["raw_fields"] = {header: values.get(column, "") for column, header in raw_headers.items() if values.get(column)}
            item["source_attachment_url"] = source_url
            item["sheet"] = sheet.title
            item["row"] = row_number
            positions.append(item)
    return positions


def parse_notice(collector: Collector, item: dict[str, str], previous_hash: str = "") -> tuple[dict[str, Any], bool]:
    response = collector.get(item["url"])
    soup = BeautifulSoup(response.text, "html.parser")
    title_node = soup.find("h1") or soup.find("h2")
    title = clean(title_node.get_text(" ", strip=True)) if title_node else item["title"]
    page_text = clean(soup.get_text(" ", strip=True))
    digest = content_hash(page_text)
    changed = digest != previous_hash

    metadata = re.search(r"日期[：:]\s*(20\d{2}-\d{2}-\d{2})\s+来源[：:]\s*([^字]+)", page_text)
    published = metadata.group(1) if metadata else item.get("published_at", "")
    publisher = clean(metadata.group(2)) if metadata else ""
    start_at, deadline = extract_period(page_text)

    attachments = []
    positions: list[dict[str, Any]] = []
    for link in soup.find_all("a", href=True):
        url = urljoin(item["url"], link["href"])
        if not ATTACHMENT_RE.search(url):
            continue
        attachment = {"name": clean(link.get_text(" ", strip=True)) or Path(url).name, "url": url}
        attachments.append(attachment)
        if re.search(r"\.xlsx(?:\?.*)?$", url, re.I):
            try:
                positions.extend(parse_workbook(collector.get(url).content, url))
                attachment["parse_status"] = "parsed"
            except Exception as exc:  # Keep the notice even when one attachment is irregular.
                attachment["parse_status"] = "failed"
                attachment["parse_error"] = clean(exc)[:240]
        else:
            attachment["parse_status"] = "stored-link-only"

    summary_match = re.search(r"本次招聘[^。]{0,120}。", page_text)
    notice = {
        "id": hashlib.sha1(item["url"].encode("utf-8")).hexdigest()[:16],
        "title": title,
        "publisher": publisher,
        "published_at": published,
        "application_start_at": start_at,
        "deadline": deadline,
        "summary": clean(summary_match.group(0)) if summary_match else "",
        "body_text": page_text,
        "source_name": "北京市人社局事业单位公开招聘",
        "source_url": item["url"],
        "content_hash": digest,
        "attachments": attachments,
        "positions": positions,
        "last_checked_at": now_iso(),
    }
    return notice, changed


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--delay-min", type=float, default=2.0)
    parser.add_argument("--delay-max", type=float, default=5.0)
    parser.add_argument("--max-items", type=int, default=50, help="Safety cap for newly discovered notices")
    parser.add_argument("--weekly", action="store_true", help="Recheck previously saved notices not on the first page")
    parser.add_argument("--reset", action="store_true", help="Ignore saved state and rebuild the local dataset")
    args = parser.parse_args()

    state = {"source": LIST_URL, "notices": {}} if args.reset else load_json(STATE_PATH, {"source": LIST_URL, "notices": {}})
    output = {"source": LIST_URL, "notices": []} if args.reset else load_json(OUTPUT_PATH, {"source": LIST_URL, "notices": []})
    saved = {notice["source_url"]: notice for notice in output.get("notices", [])}
    collector = Collector(args.delay_min, args.delay_max)

    listing = parse_listing(collector.get(LIST_URL).text)
    queue = listing[: args.max_items]
    if args.weekly:
        current = {item["url"] for item in queue}
        queue.extend(
            {"url": url, "title": saved.get(url, {}).get("title", ""), "published_at": saved.get(url, {}).get("published_at", "")}
            for url in state.get("notices", {})
            if url not in current
        )

    created = updated = unchanged = failed = 0
    for item in queue:
        url = item["url"]
        previous = state.get("notices", {}).get(url, {})
        try:
            notice, changed = parse_notice(collector, item, previous.get("content_hash", ""))
            is_new = url not in saved
            saved[url] = notice
            state.setdefault("notices", {})[url] = {
                "content_hash": notice["content_hash"],
                "last_checked_at": notice["last_checked_at"],
                "published_at": notice["published_at"],
            }
            if is_new:
                created += 1
            elif changed:
                updated += 1
            else:
                unchanged += 1
        except Exception as exc:
            failed += 1
            previous["last_error"] = clean(exc)[:400]
            previous["last_checked_at"] = now_iso()
            state.setdefault("notices", {})[url] = previous

    notices = sorted(saved.values(), key=lambda notice: (notice.get("published_at", ""), notice.get("title", "")), reverse=True)
    output = {
        "source": LIST_URL,
        "generated_at": now_iso(),
        "notice_count": len(notices),
        "position_count": sum(len(notice.get("positions", [])) for notice in notices),
        "notices": notices,
    }
    state["last_run_at"] = now_iso()
    state["last_run_summary"] = {"created": created, "updated": updated, "unchanged": unchanged, "failed": failed}
    save_json(OUTPUT_PATH, output)
    save_json(STATE_PATH, state)
    print(json.dumps(state["last_run_summary"], ensure_ascii=False))
    return 1 if failed and not (created or updated or unchanged) else 0


if __name__ == "__main__":
    raise SystemExit(main())
